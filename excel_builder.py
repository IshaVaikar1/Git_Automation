from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

FORBIDDEN = {
    "STATEMENT_OF_INCOME",
    "ITR",
    "BALANCE_SHEET",
    "PROFIT_AND_LOSS",
    "Citizens Cooperative Bank",
    "AB Bank",
    "Parner Coop Bank",
    "Aditya Bank",
    "Makarand Bank",
    "Model Test 1",
    "FSA"
}


def utc_to_ist(ts):
    if not ts:
        return ts
    try:
        dt = datetime.fromisoformat(ts)
        dt = dt + timedelta(hours=5, minutes=30)
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return ts


def style_header(row):
    for cell in row:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFC000", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")


def autosize(ws):
    for col in ws.columns:
        try:
            length = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = length + 2
        except Exception:
            continue


def contains_forbidden(values):
    for v in values:
        if not v:
            continue
        text = str(v).upper()
        for bad in FORBIDDEN:
            if bad.upper() in text:
                return True
    return False


def build_excel(payload: dict, filename="usage_report.xlsx"):

    usage_report_obj = payload.get("usageReport", {})
    rows = usage_report_obj.get("data", {}).get("data", [])

    if not isinstance(rows, list):
        raise ValueError("Invalid payload format: data.data must be a list")

    columns = [
        "bank_name",
        "tracking_id",
        "process_id",
        "file_id",
        "file_doc_type",
        "file_type",
        "file_status",
        "file_created_at",
        "file_updated_at",
        "status_msg",
        "status_code",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "UsageReport"

    ws.append(columns)
    style_header(ws[1])

    current_row = 2  # Start after header

    for req in rows:
        request_values = [
            req.get("bankName", ""),
            req.get("trackingId", "")
        ]

        if req.get("inputDocumentType") in FORBIDDEN:
            continue

        files = req.get("files", [])
        start_row = current_row

        # No files case
        if not files:
            row_values = request_values + [""] * 9
            if not contains_forbidden(row_values):
                ws.append(row_values)
                current_row += 1
            continue

        valid_file_rows = []

        for f in files:
            file_values = [
                f.get("processId", ""),
                f.get("fileId", ""),
                f.get("documentType", ""),
                f.get("fileType", ""),
                f.get("fileStatus", ""),
                utc_to_ist(f.get("createdAt", "")),
                utc_to_ist(f.get("updatedAt", "")),
                f.get("statusMessage", ""),
                f.get("statusCode", "")
            ]

            full_row = request_values + file_values

            if contains_forbidden(full_row):
                continue

            valid_file_rows.append(file_values)

        if not valid_file_rows:
            continue

        # Write rows
        for i, file_values in enumerate(valid_file_rows):
            if i == 0:
                ws.append(request_values + file_values)
            else:
                ws.append(["", ""] + file_values)
            current_row += 1

        end_row = current_row - 1

        # Merge request-level columns
        if end_row > start_row:
            for col in [1, 2]:  # bank_name, tracking_id
                ws.merge_cells(
                    start_row=start_row,
                    end_row=end_row,
                    start_column=col,
                    end_column=col
                )

                cell = ws.cell(row=start_row, column=col)
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )

    # Apply wrap text globally
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    autosize(ws)
    wb.save(filename)
    return filename